"""
Talent Review Data Analysis System
Flask application with Excel upload, data visualization, and role-based access control.
"""
import os
import json
import re
import sqlite3
import hashlib
import secrets
import io
from datetime import datetime
from functools import wraps

from flask import (Flask, request, session, redirect, url_for, jsonify,
                   render_template, g, send_from_directory, Response)
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# Regex to match illegal XML characters (not allowed in Excel cells)
_ILLEGAL_XML_CHARS_RE = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x84\x86-\x9f]')


def _sanitize_for_excel(val):
    """Remove characters that are illegal in Excel/XML worksheets."""
    if isinstance(val, str):
        return _ILLEGAL_XML_CHARS_RE.sub('', val)
    return val

from excel_parser import parse_excel_file, parse_all_excel_files, GRID_INFO, PIPELINE_ORDER

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

# Configuration - Use environment variables for cloud deployment
BASE_DIR = os.environ.get('BASE_DIR', os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA_DIR = os.environ.get('DATA_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data'))
DB_PATH = os.path.join(DATA_DIR, 'talent.db')
UPLOAD_DIR = os.path.join(DATA_DIR, 'uploads')
ALLOWED_EXTENSIONS = {'.xls', '.xlsx'}

os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ============================================================
# Database
# ============================================================

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db


@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    """Create database tables and pre-load Excel data."""
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript('''
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            sub_dept TEXT NOT NULL,
            UNIQUE(category, sub_dept)
        );

        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dept_id INTEGER NOT NULL,
            chinese_name TEXT,
            english_name TEXT,
            position_title TEXT,
            job_responsibility TEXT,
            job_level TEXT,
            age INTEGER,
            education TEXT,
            graduation_institution TEXT,
            graduation_date TEXT,
            work_experience TEXT,
            entry_date TEXT,
            company_tenure TEXT,
            base_salary REAL,
            performance_salary REAL,
            total_salary REAL,
            knowledge_skill_match REAL,
            problem_solving_match REAL,
            responsibility_match REAL,
            person_position_score REAL,
            annual_performance TEXT,
            learning_ability INTEGER,
            thinking_ability INTEGER,
            understanding_others INTEGER,
            emotional_maturity INTEGER,
            potential_score REAL,
            performance_level TEXT,
            potential_level TEXT,
            grid_position INTEGER,
            grid_name TEXT,
            talent_pipeline TEXT,
            result_application TEXT,
            development_plan TEXT,
            management_strategy TEXT,
            upload_batch_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (dept_id) REFERENCES departments(id)
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            dept_category TEXT,
            display_name TEXT
        );

        CREATE TABLE IF NOT EXISTS upload_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dept_category TEXT NOT NULL,
            sub_dept TEXT NOT NULL,
            filename TEXT,
            uploaded_by INTEGER,
            employee_count INTEGER,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_employees_dept ON employees(dept_id);
        CREATE INDEX IF NOT EXISTS idx_employees_grid ON employees(grid_position);
        CREATE INDEX IF NOT EXISTS idx_employees_pipeline ON employees(talent_pipeline);
    ''')
    db.commit()

    # Check if data already loaded
    count = db.execute('SELECT COUNT(*) as c FROM employees').fetchone()['c']
    if count == 0:
        load_excel_data(db)

    # Create default users if not exist
    create_default_users(db)
    db.close()


def load_excel_data(db):
    """Load all Excel files into database."""
    employees, depts = parse_all_excel_files(BASE_DIR)

    # Insert departments
    dept_map = {}
    for category, sub_dept in depts:
        db.execute(
            'INSERT OR IGNORE INTO departments (category, sub_dept) VALUES (?, ?)',
            (category, sub_dept)
        )
    db.commit()

    # Get dept IDs
    for row in db.execute('SELECT id, category, sub_dept FROM departments'):
        dept_map[(row['category'], row['sub_dept'])] = row['id']

    # Insert employees
    for emp in employees:
        dept_key = (emp['dept_category'], emp['sub_dept'])
        dept_id = dept_map.get(dept_key)
        if not dept_id:
            continue

        db.execute('''
            INSERT INTO employees (
                dept_id, chinese_name, english_name, position_title,
                job_responsibility, job_level, age, education,
                graduation_institution, graduation_date, work_experience,
                entry_date, company_tenure, base_salary, performance_salary,
                total_salary, knowledge_skill_match, problem_solving_match,
                responsibility_match, person_position_score, annual_performance,
                learning_ability, thinking_ability, understanding_others,
                emotional_maturity, potential_score, performance_level,
                potential_level, grid_position, grid_name, talent_pipeline,
                result_application, development_plan, management_strategy
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            dept_id, emp.get('chinese_name'), emp.get('english_name'),
            emp.get('position_title'), emp.get('job_responsibility'),
            emp.get('job_level'), emp.get('age'), emp.get('education'),
            emp.get('graduation_institution'), emp.get('graduation_date'),
            emp.get('work_experience'), emp.get('entry_date'),
            emp.get('company_tenure'), emp.get('base_salary'),
            emp.get('performance_salary'), emp.get('total_salary'),
            emp.get('knowledge_skill_match'), emp.get('problem_solving_match'),
            emp.get('responsibility_match'), emp.get('person_position_score'),
            emp.get('annual_performance'), emp.get('learning_ability'),
            emp.get('thinking_ability'), emp.get('understanding_others'),
            emp.get('emotional_maturity'), emp.get('potential_score'),
            emp.get('performance_level'), emp.get('potential_level'),
            emp.get('grid_position'), emp.get('grid_name'),
            emp.get('talent_pipeline'), emp.get('result_application'),
            emp.get('development_plan'), emp.get('management_strategy')
        ))

    # Create upload batch records
    for category, sub_dept in depts:
        emp_count = sum(1 for e in employees if e['dept_category'] == category and e['sub_dept'] == sub_dept)
        db.execute('''
            INSERT INTO upload_batches (dept_category, sub_dept, filename, uploaded_by, employee_count)
            VALUES (?, ?, ?, NULL, ?)
        ''', (category, sub_dept, f'{sub_dept}.xls', emp_count))

    db.commit()
    print(f"Loaded {len(employees)} employees from {len(depts)} departments")


def create_default_users(db):
    """Create default HR and department admin users."""
    users = [
        ('hr_admin', 'hr123456', 'hr', None, 'HR管理员'),
        ('admin_AT', 'dept123AT', 'dept_admin', '印染', '印染管理员'),
        ('admin_GB', 'dept123GB', 'dept_admin', '服装', '服装管理员'),
        ('admin_KN', 'dept123KN', 'dept_admin', '针织', '针织管理员'),
        ('admin_FL', 'dept123FL', 'dept_admin', '辅料', '辅料管理员'),
        ('admin_ZB', 'dept123ZB', 'dept_admin', '总部', '总部管理员'),
    ]
    for username, password, role, dept_cat, display_name in users:
        existing = db.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
        if not existing:
            pwd_hash = hashlib.sha256(password.encode()).hexdigest()
            db.execute(
                'INSERT INTO users (username, password_hash, role, dept_category, display_name) VALUES (?, ?, ?, ?, ?)',
                (username, pwd_hash, role, dept_cat, display_name)
            )
    db.commit()


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


# ============================================================
# Auth
# ============================================================

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': '未登录'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def hr_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'hr':
            return jsonify({'error': '权限不足，需要HR权限'}), 403
        return f(*args, **kwargs)
    return decorated


def get_current_user():
    if 'user_id' not in session:
        return None
    db = get_db()
    return db.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()


def get_accessible_dept_ids():
    """Get department IDs accessible by current user."""
    db = get_db()
    user = get_current_user()
    if not user:
        return []

    if user['role'] == 'hr':
        rows = db.execute('SELECT id FROM departments').fetchall()
    else:
        rows = db.execute(
            'SELECT id FROM departments WHERE category = ?', (user['dept_category'],)
        ).fetchall()
    return [r['id'] for r in rows]


def get_accessible_categories():
    """Get department categories accessible by current user."""
    user = get_current_user()
    if not user:
        return []
    if user['role'] == 'hr':
        return ['印染', '服装', '针织', '辅料', '总部']
    return [user['dept_category']] if user['dept_category'] else []


# ============================================================
# Routes
# ============================================================

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('role') == 'hr':
        return redirect(url_for('hr_dashboard'))
    return redirect(url_for('dept_dashboard'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        db = get_db()
        user = db.execute(
            'SELECT * FROM users WHERE username = ?', (username,)
        ).fetchone()
        if user and user['password_hash'] == hash_password(password):
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['username'] = user['username']
            session['display_name'] = user['display_name']
            session['dept_category'] = user['dept_category']
            return redirect(url_for('index'))
        return render_template('login.html', error='用户名或密码错误')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/hr-dashboard')
@login_required
@hr_required
def hr_dashboard():
    user = get_current_user()
    return render_template('hr_dashboard.html', user=user)


@app.route('/dept-dashboard')
@login_required
def dept_dashboard():
    user = get_current_user()
    return render_template('dept_dashboard.html', user=user)


@app.route('/upload')
@login_required
@hr_required
def upload_page():
    user = get_current_user()
    return render_template('upload.html', user=user)


@app.route('/export')
@login_required
def export_page():
    user = get_current_user()
    return render_template('export.html', user=user)


# ============================================================
# API Endpoints
# ============================================================

@app.route('/api/stats/overview')
@login_required
def api_overview_stats():
    db = get_db()
    dept_ids = get_accessible_dept_ids()

    if not dept_ids:
        return jsonify({'error': '无可访问部门'}), 403

    # Optional filter by category / sub_dept
    category = request.args.get('category', '')
    sub_dept = request.args.get('sub_dept', '')

    base_filter = f' AND dept_id IN ({",".join("?" * len(dept_ids))})'
    base_params = list(dept_ids)

    if category:
        base_filter += ' AND dept_id IN (SELECT id FROM departments WHERE category = ?)'
        base_params.append(category)
    if sub_dept:
        base_filter += ' AND dept_id IN (SELECT id FROM departments WHERE sub_dept = ?)'
        base_params.append(sub_dept)

    total = db.execute(
        'SELECT COUNT(*) as c FROM employees WHERE 1=1' + base_filter, base_params
    ).fetchone()['c']

    dept_count = db.execute(
        'SELECT COUNT(DISTINCT dept_id) as c FROM employees WHERE 1=1' + base_filter, base_params
    ).fetchone()['c']

    avg_age = db.execute(
        'SELECT AVG(age) as v FROM employees WHERE 1=1' + base_filter + ' AND age IS NOT NULL', base_params
    ).fetchone()['v']

    avg_tenure_months = db.execute(f"""
        SELECT AVG(CAST(
            CASE
                WHEN company_tenure LIKE '%年%个月' THEN
                    CAST(SUBSTR(company_tenure, 1, INSTR(company_tenure, '年')-1) AS INTEGER) * 12 +
                    CAST(SUBSTR(company_tenure, INSTR(company_tenure, '年')+1, INSTR(company_tenure, '个月')-INSTR(company_tenure, '年')-1) AS INTEGER)
                WHEN company_tenure LIKE '%年' THEN
                    CAST(SUBSTR(company_tenure, 1, INSTR(company_tenure, '年')-1) AS INTEGER) * 12
                WHEN company_tenure LIKE '%个月' THEN
                    CAST(SUBSTR(company_tenure, 1, INSTR(company_tenure, '个月')-1) AS INTEGER)
                ELSE 0
            END AS REAL)) as v
        FROM employees WHERE 1=1""" + base_filter + " AND company_tenure != ''", base_params
    ).fetchone()['v']

    # Pipeline distribution
    pipeline_rows = db.execute("""
        SELECT talent_pipeline, COUNT(*) as c
        FROM employees WHERE 1=1""" + base_filter + """
        GROUP BY talent_pipeline
    """, base_params).fetchall()
    pipeline_dist = {r['talent_pipeline']: r['c'] for r in pipeline_rows}

    # Grid distribution
    grid_rows = db.execute("""
        SELECT grid_position, COUNT(*) as c
        FROM employees WHERE 1=1""" + base_filter + " AND grid_position IS NOT NULL" + """
        GROUP BY grid_position
    """, base_params).fetchall()
    grid_dist = {str(r['grid_position']): r['c'] for r in grid_rows}

    # Performance distribution
    perf_rows = db.execute("""
        SELECT annual_performance, COUNT(*) as c
        FROM employees WHERE 1=1""" + base_filter + " AND annual_performance != ''" + """
        GROUP BY annual_performance
    """, base_params).fetchall()
    perf_dist = {r['annual_performance']: r['c'] for r in perf_rows}

    # Education distribution
    edu_rows = db.execute("""
        SELECT education, COUNT(*) as c
        FROM employees WHERE 1=1""" + base_filter + " AND education != ''" + """
        GROUP BY education ORDER BY c DESC
    """, base_params).fetchall()
    edu_dist = {r['education']: r['c'] for r in edu_rows}

    return jsonify({
        'total_employees': total,
        'dept_count': dept_count,
        'avg_age': round(avg_age, 1) if avg_age else 0,
        'avg_tenure_months': round(avg_tenure_months, 1) if avg_tenure_months else 0,
        'pipeline_dist': pipeline_dist,
        'grid_dist': grid_dist,
        'perf_dist': perf_dist,
        'edu_dist': edu_dist,
    })


@app.route('/api/stats/department-comparison')
@login_required
def api_dept_comparison():
    db = get_db()
    dept_ids = get_accessible_dept_ids()
    if not dept_ids:
        return jsonify([])

    placeholders = ','.join('?' * len(dept_ids))
    rows = db.execute(f"""
        SELECT d.category, d.sub_dept, COUNT(e.id) as emp_count,
               AVG(e.age) as avg_age,
               AVG(e.person_position_score) as avg_match
        FROM departments d
        LEFT JOIN employees e ON e.dept_id = d.id
        WHERE d.id IN ({placeholders})
        GROUP BY d.id
        HAVING emp_count > 0
        ORDER BY d.category, d.sub_dept
    """, dept_ids).fetchall()

    result = []
    for r in rows:
        result.append({
            'category': r['category'],
            'sub_dept': r['sub_dept'],
            'emp_count': r['emp_count'],
            'avg_age': round(r['avg_age'], 1) if r['avg_age'] else 0,
            'avg_match': round(r['avg_match'], 2) if r['avg_match'] else 0,
        })
    return jsonify(result)


@app.route('/api/stats/category-summary')
@login_required
def api_category_summary():
    """Summary by major department category."""
    db = get_db()
    categories = get_accessible_categories()

    result = []
    for cat in categories:
        rows = db.execute("""
            SELECT e.grid_position, e.talent_pipeline, e.annual_performance,
                   e.age, e.person_position_score
            FROM employees e
            JOIN departments d ON e.dept_id = d.id
            WHERE d.category = ?
        """, (cat,)).fetchall()

        if not rows:
            continue

        total = len(rows)
        grid_dist = {}
        pipeline_dist = {}
        perf_dist = {}
        ages = []
        match_scores = []

        for r in rows:
            if r['grid_position']:
                grid_dist[r['grid_position']] = grid_dist.get(r['grid_position'], 0) + 1
            if r['talent_pipeline']:
                pipeline_dist[r['talent_pipeline']] = pipeline_dist.get(r['talent_pipeline'], 0) + 1
            if r['annual_performance']:
                perf_dist[r['annual_performance']] = perf_dist.get(r['annual_performance'], 0) + 1
            if r['age']:
                ages.append(r['age'])
            if r['person_position_score']:
                match_scores.append(r['person_position_score'])

        result.append({
            'category': cat,
            'total': total,
            'avg_age': round(sum(ages)/len(ages), 1) if ages else 0,
            'avg_match': round(sum(match_scores)/len(match_scores), 2) if match_scores else 0,
            'grid_dist': {str(k): v for k, v in sorted(grid_dist.items())},
            'pipeline_dist': pipeline_dist,
            'perf_dist': perf_dist,
        })

    return jsonify(result)


@app.route('/api/stats/grid')
@login_required
def api_grid_stats():
    """9-box grid statistics with employee names."""
    db = get_db()
    dept_ids = get_accessible_dept_ids()
    if not dept_ids:
        return jsonify({})

    # Optional filter by category
    category = request.args.get('category', '')
    sub_dept = request.args.get('sub_dept', '')

    query = f"""
        SELECT e.id, e.chinese_name, e.position_title, e.grid_position,
               e.performance_level, e.potential_level, e.talent_pipeline,
               d.category, d.sub_dept
        FROM employees e
        JOIN departments d ON e.dept_id = d.id
        WHERE e.dept_id IN ({','.join('?' * len(dept_ids))})
    """
    params = list(dept_ids)

    if category:
        query += ' AND d.category = ?'
        params.append(category)
    if sub_dept:
        query += ' AND d.sub_dept = ?'
        params.append(sub_dept)

    rows = db.execute(query, params).fetchall()

    grid_data = {}
    for grid_pos in range(1, 10):
        info = GRID_INFO.get(grid_pos, {})
        grid_data[str(grid_pos)] = {
            'name': info.get('name', ''),
            'perf': info.get('perf', ''),
            'pot': info.get('pot', ''),
            'color': info.get('color', '#d9d9d9'),
            'count': 0,
            'employees': []
        }

    for r in rows:
        gp = r['grid_position']
        if gp and 1 <= gp <= 9:
            cell = grid_data[str(gp)]
            cell['count'] += 1
            cell['employees'].append({
                'id': r['id'],
                'name': r['chinese_name'],
                'position': r['position_title'],
                'category': r['category'],
                'sub_dept': r['sub_dept'],
                'pipeline': r['talent_pipeline'],
            })

    return jsonify(grid_data)


@app.route('/api/stats/age-distribution')
@login_required
def api_age_dist():
    db = get_db()
    dept_ids = get_accessible_dept_ids()
    if not dept_ids:
        return jsonify([])

    category = request.args.get('category', '')
    sub_dept = request.args.get('sub_dept', '')

    base_filter = f' AND dept_id IN ({",".join("?" * len(dept_ids))})'
    base_params = list(dept_ids)

    if category:
        base_filter += ' AND dept_id IN (SELECT id FROM departments WHERE category = ?)'
        base_params.append(category)
    if sub_dept:
        base_filter += ' AND dept_id IN (SELECT id FROM departments WHERE sub_dept = ?)'
        base_params.append(sub_dept)

    rows = db.execute(f"""
        SELECT
            CASE
                WHEN age < 25 THEN '25岁以下'
                WHEN age >= 25 AND age < 30 THEN '25-29岁'
                WHEN age >= 30 AND age < 35 THEN '30-34岁'
                WHEN age >= 35 AND age < 40 THEN '35-39岁'
                WHEN age >= 40 AND age < 45 THEN '40-44岁'
                WHEN age >= 45 AND age < 50 THEN '45-49岁'
                WHEN age >= 50 THEN '50岁以上'
            END as age_group,
            COUNT(*) as count
        FROM employees
        WHERE 1=1""" + base_filter + " AND age IS NOT NULL" + """
        GROUP BY age_group
        ORDER BY age_group
    """, base_params).fetchall()

    return jsonify([{'label': r['age_group'], 'count': r['count']} for r in rows])


@app.route('/api/stats/pipeline')
@login_required
def api_pipeline_stats():
    """Talent pipeline distribution by category."""
    db = get_db()
    categories = get_accessible_categories()
    category = request.args.get('category', '')

    if category and category in categories:
        cats = [category]
    else:
        cats = categories

    result = {}
    for cat in cats:
        rows = db.execute("""
            SELECT talent_pipeline, COUNT(*) as c
            FROM employees e
            JOIN departments d ON e.dept_id = d.id
            WHERE d.category = ? AND talent_pipeline != ''
            GROUP BY talent_pipeline
        """, (cat,)).fetchall()
        result[cat] = {r['talent_pipeline']: r['c'] for r in rows}

    return jsonify(result)


@app.route('/api/employees')
@login_required
def api_employees():
    """List employees with optional filters."""
    db = get_db()
    dept_ids = get_accessible_dept_ids()
    if not dept_ids:
        return jsonify({'employees': [], 'total': 0})

    category = request.args.get('category', '')
    sub_dept = request.args.get('sub_dept', '')
    grid_pos = request.args.get('grid', '')
    pipeline = request.args.get('pipeline', '')
    search = request.args.get('search', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))

    query = f"""
        SELECT e.*, d.category, d.sub_dept
        FROM employees e
        JOIN departments d ON e.dept_id = d.id
        WHERE e.dept_id IN ({','.join('?' * len(dept_ids))})
    """
    params = list(dept_ids)
    count_query = f"""
        SELECT COUNT(*) as c
        FROM employees e
        JOIN departments d ON e.dept_id = d.id
        WHERE e.dept_id IN ({','.join('?' * len(dept_ids))})
    """
    count_params = list(dept_ids)

    if category:
        query += ' AND d.category = ?'
        count_query += ' AND d.category = ?'
        params.append(category)
        count_params.append(category)
    if sub_dept:
        query += ' AND d.sub_dept = ?'
        count_query += ' AND d.sub_dept = ?'
        params.append(sub_dept)
        count_params.append(sub_dept)
    if grid_pos:
        query += ' AND e.grid_position = ?'
        count_query += ' AND e.grid_position = ?'
        params.append(int(grid_pos))
        count_params.append(int(grid_pos))
    if pipeline:
        query += ' AND e.talent_pipeline = ?'
        count_query += ' AND e.talent_pipeline = ?'
        params.append(pipeline)
        count_params.append(pipeline)
    if search:
        query += ' AND (e.chinese_name LIKE ? OR e.position_title LIKE ?)'
        count_query += ' AND (e.chinese_name LIKE ? OR e.position_title LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
        count_params.extend([f'%{search}%', f'%{search}%'])

    total = db.execute(count_query, count_params).fetchone()['c']

    query += ' ORDER BY d.category, d.sub_dept, e.id LIMIT ? OFFSET ?'
    params.extend([per_page, (page - 1) * per_page])

    rows = db.execute(query, params).fetchall()

    employees = []
    for r in rows:
        emp = dict(r)
        # Convert None values to empty strings for JSON
        for k, v in emp.items():
            if v is None:
                emp[k] = ''
        employees.append(emp)

    return jsonify({'employees': employees, 'total': total, 'page': page, 'per_page': per_page})


@app.route('/api/employee/<int:emp_id>')
@login_required
def api_employee_detail(emp_id):
    db = get_db()
    dept_ids = get_accessible_dept_ids()

    row = db.execute("""
        SELECT e.*, d.category, d.sub_dept
        FROM employees e
        JOIN departments d ON e.dept_id = d.id
        WHERE e.id = ? AND e.dept_id IN (%s)
    """ % ','.join('?' * len(dept_ids)), [emp_id] + dept_ids).fetchone()

    if not row:
        return jsonify({'error': '员工不存在或无权访问'}), 404

    emp = dict(row)
    for k, v in emp.items():
        if v is None:
            emp[k] = ''

    # Add grid info
    grid_pos = emp.get('grid_position')
    if grid_pos and grid_pos in GRID_INFO:
        info = GRID_INFO[grid_pos]
        emp['grid_info'] = info
    else:
        emp['grid_info'] = None

    return jsonify(emp)


@app.route('/api/departments')
@login_required
def api_departments():
    db = get_db()
    dept_ids = get_accessible_dept_ids()
    placeholders = ','.join('?' * len(dept_ids))

    rows = db.execute(f"""
        SELECT d.id, d.category, d.sub_dept, COUNT(e.id) as emp_count
        FROM departments d
        LEFT JOIN employees e ON e.dept_id = d.id
        WHERE d.id IN ({placeholders})
        GROUP BY d.id
        ORDER BY d.category, d.sub_dept
    """, dept_ids).fetchall()

    result = {}
    for r in rows:
        cat = r['category']
        if cat not in result:
            result[cat] = []
        result[cat].append({
            'id': r['id'],
            'sub_dept': r['sub_dept'],
            'emp_count': r['emp_count']
        })

    return jsonify(result)


@app.route('/api/departments/add', methods=['POST'])
@login_required
@hr_required
def api_add_department():
    """Add a new department (HR only)."""
    db = get_db()
    category = request.json.get('category', '').strip() if request.is_json else request.form.get('category', '').strip()
    sub_dept = request.json.get('sub_dept', '').strip() if request.is_json else request.form.get('sub_dept', '').strip()

    if not category or not sub_dept:
        return jsonify({'error': '部门大类和子部门名称不能为空'}), 400

    existing = db.execute(
        'SELECT id FROM departments WHERE category = ? AND sub_dept = ?',
        (category, sub_dept)
    ).fetchone()

    if existing:
        return jsonify({'error': f'部门已存在: {category}/{sub_dept}'}), 409

    db.execute(
        'INSERT INTO departments (category, sub_dept) VALUES (?, ?)',
        (category, sub_dept)
    )
    db.commit()

    return jsonify({'success': True, 'message': f'部门已添加: {category}/{sub_dept}'})


@app.route('/api/departments/delete', methods=['POST'])
@login_required
@hr_required
def api_delete_department():
    """Delete a department and all its employees (HR only)."""
    db = get_db()
    dept_id = request.json.get('dept_id') if request.is_json else request.form.get('dept_id')

    if not dept_id:
        return jsonify({'error': '缺少部门ID'}), 400

    dept = db.execute('SELECT * FROM departments WHERE id = ?', (dept_id,)).fetchone()
    if not dept:
        return jsonify({'error': '部门不存在'}), 404

    emp_count = db.execute('SELECT COUNT(*) as c FROM employees WHERE dept_id = ?', (dept_id,)).fetchone()['c']

    db.execute('DELETE FROM employees WHERE dept_id = ?', (dept_id,))
    db.execute('DELETE FROM departments WHERE id = ?', (dept_id,))
    db.commit()

    return jsonify({
        'success': True,
        'message': f'已删除部门: {dept["category"]}/{dept["sub_dept"]}（含{emp_count}名员工）'
    })


@app.route('/api/upload/history')
@login_required
def api_upload_history():
    db = get_db()
    user = get_current_user()

    if user['role'] == 'hr':
        rows = db.execute("""
            SELECT b.*, u.display_name as uploader
            FROM upload_batches b
            LEFT JOIN users u ON b.uploaded_by = u.id
            ORDER BY b.uploaded_at DESC
        """).fetchall()
    else:
        rows = db.execute("""
            SELECT b.*, u.display_name as uploader
            FROM upload_batches b
            LEFT JOIN users u ON b.uploaded_by = u.id
            WHERE b.dept_category = ?
            ORDER BY b.uploaded_at DESC
        """, (user['dept_category'],)).fetchall()

    return jsonify([dict(r) for r in rows])


@app.route('/api/upload', methods=['POST'])
@login_required
@hr_required
def api_upload():
    user = get_current_user()
    if not user:
        return jsonify({'error': '未登录'}), 401

    if 'file' not in request.files:
        return jsonify({'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'error': f'不支持的文件格式: {ext}，请上传 .xls 或 .xlsx 文件'}), 400

    # Determine department
    category = request.form.get('category', '')
    sub_dept = request.form.get('sub_dept', '')

    # Department admin can only upload to their own category
    if user['role'] == 'dept_admin':
        if category and category != user['dept_category']:
            return jsonify({'error': '无权上传到其他部门'}), 403
        category = user['dept_category']

    if not category:
        return jsonify({'error': '请选择部门大类'}), 400
    if not sub_dept:
        # Try to extract from filename
        sub_dept = os.path.splitext(secure_filename(file.filename))[0]

    # Save file
    filename = secure_filename(file.filename)
    if not filename:
        filename = f'upload_{datetime.now().strftime("%Y%m%d_%H%M%S")}{ext}'
    filepath = os.path.join(UPLOAD_DIR, f'{category}_{sub_dept}_{filename}')
    file.save(filepath)

    # Parse Excel
    try:
        employees = parse_excel_file(filepath, category, sub_dept)
    except Exception as e:
        os.remove(filepath)
        return jsonify({'error': f'Excel解析失败: {str(e)}'}), 400

    if not employees:
        os.remove(filepath)
        return jsonify({'error': '文件中未找到有效数据，请检查Excel格式是否符合人才九宫格模板'}), 400

    # Save to database
    db = get_db()

    # Get or create department
    dept = db.execute(
        'SELECT id FROM departments WHERE category = ? AND sub_dept = ?',
        (category, sub_dept)
    ).fetchone()

    if dept:
        dept_id = dept['id']
        # Delete old employees for this department
        db.execute('DELETE FROM employees WHERE dept_id = ?', (dept_id,))
    else:
        cur = db.execute(
            'INSERT INTO departments (category, sub_dept) VALUES (?, ?)',
            (category, sub_dept)
        )
        dept_id = cur.lastrowid

    # Insert new employees
    for emp in employees:
        db.execute('''
            INSERT INTO employees (
                dept_id, chinese_name, english_name, position_title,
                job_responsibility, job_level, age, education,
                graduation_institution, graduation_date, work_experience,
                entry_date, company_tenure, base_salary, performance_salary,
                total_salary, knowledge_skill_match, problem_solving_match,
                responsibility_match, person_position_score, annual_performance,
                learning_ability, thinking_ability, understanding_others,
                emotional_maturity, potential_score, performance_level,
                potential_level, grid_position, grid_name, talent_pipeline,
                result_application, development_plan, management_strategy,
                upload_batch_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            dept_id, emp.get('chinese_name'), emp.get('english_name'),
            emp.get('position_title'), emp.get('job_responsibility'),
            emp.get('job_level'), emp.get('age'), emp.get('education'),
            emp.get('graduation_institution'), emp.get('graduation_date'),
            emp.get('work_experience'), emp.get('entry_date'),
            emp.get('company_tenure'), emp.get('base_salary'),
            emp.get('performance_salary'), emp.get('total_salary'),
            emp.get('knowledge_skill_match'), emp.get('problem_solving_match'),
            emp.get('responsibility_match'), emp.get('person_position_score'),
            emp.get('annual_performance'), emp.get('learning_ability'),
            emp.get('thinking_ability'), emp.get('understanding_others'),
            emp.get('emotional_maturity'), emp.get('potential_score'),
            emp.get('performance_level'), emp.get('potential_level'),
            emp.get('grid_position'), emp.get('grid_name'),
            emp.get('talent_pipeline'), emp.get('result_application'),
            emp.get('development_plan'), emp.get('management_strategy'), None
        ))

    # Record upload batch
    batch_cur = db.execute('''
        INSERT INTO upload_batches (dept_category, sub_dept, filename, uploaded_by, employee_count)
        VALUES (?, ?, ?, ?, ?)
    ''', (category, sub_dept, filename, user['id'], len(employees)))
    batch_id = batch_cur.lastrowid

    # Update employees with batch_id
    db.execute('UPDATE employees SET upload_batch_id = ? WHERE dept_id = ?', (batch_id, dept_id))

    db.commit()

    return jsonify({
        'success': True,
        'message': f'成功导入 {len(employees)} 名员工数据',
        'employee_count': len(employees),
        'category': category,
        'sub_dept': sub_dept
    })


@app.route('/api/grid-info')
@login_required
def api_grid_info():
    """Return 9-box grid configuration."""
    return jsonify({str(k): v for k, v in GRID_INFO.items()})


@app.route('/api/stats/analysis')
@login_required
def api_analysis():
    """Generate intelligent analysis and recommendations based on filtered data."""
    db = get_db()
    dept_ids = get_accessible_dept_ids()
    if not dept_ids:
        return jsonify({'error': '无可访问部门'}), 403

    category = request.args.get('category', '')
    sub_dept = request.args.get('sub_dept', '')

    base_filter = f' AND e.dept_id IN ({",".join("?" * len(dept_ids))})'
    base_params = list(dept_ids)

    if category:
        base_filter += ' AND d.category = ?'
        base_params.append(category)
    if sub_dept:
        base_filter += ' AND d.sub_dept = ?'
        base_params.append(sub_dept)

    # Fetch aggregated data
    rows = db.execute(f"""
        SELECT e.age, e.education, e.annual_performance, e.grid_position,
               e.talent_pipeline, e.person_position_score, e.potential_score,
               e.company_tenure, e.performance_level, e.potential_level
        FROM employees e
        JOIN departments d ON e.dept_id = d.id
        WHERE 1=1""" + base_filter, base_params
    ).fetchall()

    if not rows:
        return jsonify({'error': '无数据'}), 404

    total = len(rows)
    ages = [r['age'] for r in rows if r['age']]
    avg_age = round(sum(ages) / len(ages), 1) if ages else 0

    # Grid distribution
    grid_counts = {}
    for r in rows:
        gp = r['grid_position']
        if gp:
            grid_counts[gp] = grid_counts.get(gp, 0) + 1

    star_count = grid_counts.get(9, 0)
    perf_star = grid_counts.get(8, 0)
    pot_star = grid_counts.get(7, 0)
    core = grid_counts.get(5, 0)
    problem = grid_counts.get(1, 0)
    gap = grid_counts.get(2, 0)
    high_pot = star_count + pot_star + grid_counts.get(4, 0)
    high_perf = star_count + perf_star + grid_counts.get(6, 0)

    # Pipeline distribution
    pipeline_counts = {}
    for r in rows:
        p = r['talent_pipeline']
        if p:
            pipeline_counts[p] = pipeline_counts.get(p, 0) + 1

    # Performance distribution
    perf_counts = {}
    for r in rows:
        p = r['annual_performance']
        if p:
            perf_counts[p] = perf_counts.get(p, 0) + 1

    # Education distribution
    edu_counts = {}
    for r in rows:
        e = r['education']
        if e:
            edu_counts[e] = edu_counts.get(e, 0) + 1

    # Age distribution
    age_groups = {'25岁以下': 0, '25-29岁': 0, '30-34岁': 0, '35-39岁': 0, '40-44岁': 0, '45-49岁': 0, '50岁以上': 0}
    for a in ages:
        if a < 25: age_groups['25岁以下'] += 1
        elif a < 30: age_groups['25-29岁'] += 1
        elif a < 35: age_groups['30-34岁'] += 1
        elif a < 40: age_groups['35-39岁'] += 1
        elif a < 45: age_groups['40-44岁'] += 1
        elif a < 50: age_groups['45-49岁'] += 1
        else: age_groups['50岁以上'] += 1

    # Match scores
    match_scores = [r['person_position_score'] for r in rows if r['person_position_score']]
    avg_match = round(sum(match_scores) / len(match_scores), 2) if match_scores else 0

    # Tenure
    tenures = []
    for r in rows:
        t = r['company_tenure']
        if t and t != '':
            months = 0
            if '年' in t and '个月' in t:
                parts = t.split('年')
                months = int(parts[0]) * 12 if parts[0].strip() else 0
                rest = parts[1].replace('个月', '').strip()
                months += int(rest) if rest else 0
            elif '年' in t:
                months = int(t.replace('年', '').strip()) * 12
            elif '个月' in t:
                months = int(t.replace('个月', '').strip())
            if months > 0:
                tenures.append(months)
    avg_tenure = round(sum(tenures) / len(tenures), 1) if tenures else 0

    # ---- Generate analysis text ----
    analyses = []

    # 1. Overall summary
    scope = ''
    if category: scope = category
    if sub_dept: scope += f' / {sub_dept}'
    if not scope: scope = '全部可访问部门'

    analyses.append({
        'title': '总体概况',
        'type': 'info',
        'content': f'当前分析范围：{scope}，共 {total} 人。平均年龄 {avg_age} 岁，平均司龄 {avg_tenure} 个月，平均人岗匹配评分 {avg_match} 分。'
    })

    # 2. Talent structure
    high_pot_pct = round(high_pot / total * 100, 1) if total else 0
    high_perf_pct = round(high_perf / total * 100, 1) if total else 0
    star_pct = round(star_count / total * 100, 1) if total else 0
    problem_pct = round((problem + gap) / total * 100, 1) if total else 0

    structure_text = f'高潜人才 {high_pot} 人（占比 {high_pot_pct}%），高绩效人才 {high_perf} 人（占比 {high_perf_pct}%），超级明星 {star_count} 人（{star_pct}%），中坚力量 {core} 人。'
    if problem_pct > 15:
        structure_text += f' ⚠️ 问题员工和差距员工合计 {(problem+gap)} 人（{problem_pct}%），占比偏高，需重点关注绩效改进。'
        analyses.append({'title': '人才结构分析', 'type': 'warning', 'content': structure_text})
    elif star_pct > 15:
        structure_text += f' ★ 超级明星占比 {star_pct}%，人才质量优秀，建议加强梯队建设与保留机制。'
        analyses.append({'title': '人才结构分析', 'type': 'success', 'content': structure_text})
    else:
        analyses.append({'title': '人才结构分析', 'type': 'info', 'content': structure_text})

    # 3. Pipeline analysis
    t1 = pipeline_counts.get('第一梯队', 0)
    t2 = pipeline_counts.get('第二梯队', 0)
    t3 = pipeline_counts.get('第三梯队', 0)
    t4 = pipeline_counts.get('第四梯队', 0)
    t5 = pipeline_counts.get('第五梯队', 0)
    others = pipeline_counts.get('其他', 0)
    pipeline_text = f'第一梯队（关键人才储备）{t1} 人，第二梯队 {t2} 人，第三梯队 {t3} 人，第四梯队 {t4} 人，第五梯队 {t5} 人，其他 {others} 人。'
    if t1 + t2 < total * 0.1:
        pipeline_text += f' ⚠️ 第一、二梯队合计 {t1+t2} 人，仅占 {round((t1+t2)/total*100,1)}%，关键人才储备不足，需加快梯队建设。'
        analyses.append({'title': '人才梯队分析', 'type': 'warning', 'content': pipeline_text})
    else:
        pipeline_text += f' 第一、二梯队合计 {t1+t2} 人，占比 {round((t1+t2)/total*100,1)}%，梯队结构较为合理。'
        analyses.append({'title': '人才梯队分析', 'type': 'info', 'content': pipeline_text})

    # 4. Performance analysis
    a_count = perf_counts.get('A', 0)
    b_count = perf_counts.get('B', 0)
    c_count = perf_counts.get('C', 0)
    a_pct = round(a_count / total * 100, 1) if total else 0
    c_pct = round(c_count / total * 100, 1) if total else 0
    perf_text = f'A（优秀）{a_count} 人，B（达标）{b_count} 人，C（待改进）{c_count} 人。A类占比 {a_pct}%。'
    if c_pct > 15:
        perf_text += f' ⚠️ C类员工占比 {c_pct}%，绩效不达标人数较多，建议制定绩效改进计划（PIP）。'
        analyses.append({'title': '绩效表现分析', 'type': 'warning', 'content': perf_text})
    elif a_pct > 30:
        perf_text += f' ★ A类占比 {a_pct}%，整体绩效表现优秀。'
        analyses.append({'title': '绩效表现分析', 'type': 'success', 'content': perf_text})
    else:
        analyses.append({'title': '绩效表现分析', 'type': 'info', 'content': perf_text})

    # 5. Age structure
    young = age_groups['25岁以下'] + age_groups['25-29岁']
    middle = age_groups['30-34岁'] + age_groups['35-39岁']
    senior = age_groups['40-44岁'] + age_groups['45-49岁'] + age_groups['50岁以上']
    young_pct = round(young / total * 100, 1) if total else 0
    senior_pct = round(senior / total * 100, 1) if total else 0
    age_text = f'30岁以下 {young} 人（{young_pct}%），30-39岁 {middle} 人，40岁以上 {senior} 人（{senior_pct}%）。'
    if senior_pct > 50:
        age_text += f' ⚠️ 40岁以上员工占比 {senior_pct}%，年龄结构偏大，需加强年轻人才引进与梯队补充。'
        analyses.append({'title': '年龄结构分析', 'type': 'warning', 'content': age_text})
    elif young_pct > 35:
        age_text += f' ★ 30岁以下员工占比 {young_pct}%，团队年轻化，富有活力，需注重经验传承与培养。'
        analyses.append({'title': '年龄结构分析', 'type': 'success', 'content': age_text})
    else:
        analyses.append({'title': '年龄结构分析', 'type': 'info', 'content': age_text})

    # 6. Education analysis
    bachelor_plus = edu_counts.get('本科', 0) + edu_counts.get('硕士研究生', 0) + edu_counts.get('硕士', 0) + edu_counts.get('博士', 0)
    high_school = edu_counts.get('高中/高职/中专', 0) + edu_counts.get('高中', 0)
    middle_school = edu_counts.get('初中', 0) + edu_counts.get('小学', 0)
    college = edu_counts.get('大专', 0)
    edu_pct = round(bachelor_plus / total * 100, 1) if total else 0
    low_edu_pct = round((high_school + middle_school) / total * 100, 1) if total else 0
    edu_text = f'本科及以上 {bachelor_plus} 人（{edu_pct}%），大专 {college} 人，高中及以下 {high_school + middle_school} 人（{low_edu_pct}%）。'
    if edu_pct < 20:
        edu_text += f' ⚠️ 本科及以上学历占比偏低（{edu_pct}%），建议针对核心岗位提升学历要求或加强专业培训。'
        analyses.append({'title': '学历结构分析', 'type': 'warning', 'content': edu_text})
    else:
        edu_text += f' 学历结构整体{edu_pct > 40 and "较好" or "适中"}。'
        analyses.append({'title': '学历结构分析', 'type': 'info', 'content': edu_text})

    # 7. Recommendations
    recommendations = []
    if star_count > 0:
        recommendations.append(f'★ 超级明星员工 {star_count} 人是核心资产，建议制定个性化保留方案（晋升、轮岗、关键项目历练），防止人才流失。')
    if high_pot > 0:
        recommendations.append(f'★ {high_pot} 名高潜人才需重点培养，建议安排导师辅导、跨部门轮岗、领导力培训等发展项目。')
    if problem + gap > 0:
        recommendations.append(f'⚠️ {problem+gap} 名问题/差距员工需启动绩效辅导，设定3-6个月改进期限，无明显改善则启动退出机制。')
    if c_count > 0:
        recommendations.append(f'⚠️ {c_count} 名C类绩效员工需一对一沟通，明确改进目标，必要时调整岗位或职责。')
    if senior_pct > 50:
        recommendations.append(f'⚠️ 团队年龄结构偏大，建议加大应届生/年轻人才招聘力度，做好知识沉淀与经验传承。')
    if t1 + t2 < total * 0.1:
        recommendations.append(f'⚠️ 关键梯队人才储备不足，建议从高潜员工中选拔建立后备人才库，制定加速发展计划。')
    if avg_match < 7:
        recommendations.append(f'⚠️ 平均人岗匹配评分 {avg_match} 分偏低，建议重新审视岗位配置与人岗匹配度，优化人员安排。')
    if not recommendations:
        recommendations.append('当前部门人才结构整体健康，建议持续关注人才发展与梯队建设，保持定期盘点与动态调整。')

    analyses.append({
        'title': '管理建议',
        'type': 'recommendation',
        'content': '\n'.join(recommendations)
    })

    return jsonify({
        'total': total,
        'avg_age': avg_age,
        'avg_match': avg_match,
        'avg_tenure': avg_tenure,
        'grid_counts': {str(k): v for k, v in grid_counts.items()},
        'pipeline_counts': pipeline_counts,
        'perf_counts': perf_counts,
        'edu_counts': edu_counts,
        'age_groups': age_groups,
        'high_pot': high_pot,
        'high_perf': high_perf,
        'star_count': star_count,
        'analyses': analyses,
    })


def _generate_analysis_text(total, scope, avg_age, avg_tenure, avg_match,
                            grid_counts, pipeline_counts, perf_counts,
                            edu_counts, age_groups, star_count, high_pot,
                            high_perf, core_count, problem_count):
    """Generate intelligent analysis text for export. Returns list of dicts with title/type/content."""
    analyses = []

    # 1. Overall summary
    analyses.append({
        'title': '总体概况',
        'type': 'info',
        'content': f'当前分析范围：{scope}，共 {total} 人。平均年龄 {avg_age} 岁，平均司龄 {avg_tenure} 个月，平均人岗匹配评分 {avg_match} 分。'
    })

    # 2. Talent structure
    high_pot_pct = round(high_pot / total * 100, 1) if total else 0
    high_perf_pct = round(high_perf / total * 100, 1) if total else 0
    star_pct = round(star_count / total * 100, 1) if total else 0
    problem_pct = round(problem_count / total * 100, 1) if total else 0

    structure_text = f'高潜人才 {high_pot} 人（占比 {high_pot_pct}%），高绩效人才 {high_perf} 人（占比 {high_perf_pct}%），超级明星 {star_count} 人（{star_pct}%），中坚力量 {core_count} 人。'
    if problem_pct > 15:
        structure_text += f' 问题员工和差距员工合计 {problem_count} 人（{problem_pct}%），占比偏高，需重点关注绩效改进。'
        analyses.append({'title': '人才结构分析', 'type': 'warning', 'content': structure_text})
    elif star_pct > 15:
        structure_text += f' 超级明星占比 {star_pct}%，人才质量优秀，建议加强梯队建设与保留机制。'
        analyses.append({'title': '人才结构分析', 'type': 'success', 'content': structure_text})
    else:
        analyses.append({'title': '人才结构分析', 'type': 'info', 'content': structure_text})

    # 3. Pipeline analysis
    t1 = pipeline_counts.get('第一梯队', 0)
    t2 = pipeline_counts.get('第二梯队', 0)
    t3 = pipeline_counts.get('第三梯队', 0)
    t4 = pipeline_counts.get('第四梯队', 0)
    t5 = pipeline_counts.get('第五梯队', 0)
    others = pipeline_counts.get('其他', 0)
    pipeline_text = f'第一梯队（关键人才储备）{t1} 人，第二梯队 {t2} 人，第三梯队 {t3} 人，第四梯队 {t4} 人，第五梯队 {t5} 人，其他 {others} 人。'
    if t1 + t2 < total * 0.1:
        pipeline_text += f' 第一、二梯队合计 {t1+t2} 人，仅占 {round((t1+t2)/total*100,1)}%，关键人才储备不足，需加快梯队建设。'
        analyses.append({'title': '人才梯队分析', 'type': 'warning', 'content': pipeline_text})
    else:
        pipeline_text += f' 第一、二梯队合计 {t1+t2} 人，占比 {round((t1+t2)/total*100,1)}%，梯队结构较为合理。'
        analyses.append({'title': '人才梯队分析', 'type': 'info', 'content': pipeline_text})

    # 4. Performance analysis
    a_count = perf_counts.get('A', 0)
    b_count = perf_counts.get('B', 0)
    c_count = perf_counts.get('C', 0)
    a_pct = round(a_count / total * 100, 1) if total else 0
    c_pct = round(c_count / total * 100, 1) if total else 0
    perf_text = f'A（优秀）{a_count} 人，B（达标）{b_count} 人，C（待改进）{c_count} 人。A类占比 {a_pct}%。'
    if c_pct > 15:
        perf_text += f' C类员工占比 {c_pct}%，绩效不达标人数较多，建议制定绩效改进计划（PIP）。'
        analyses.append({'title': '绩效表现分析', 'type': 'warning', 'content': perf_text})
    elif a_pct > 30:
        perf_text += f' A类占比 {a_pct}%，整体绩效表现优秀。'
        analyses.append({'title': '绩效表现分析', 'type': 'success', 'content': perf_text})
    else:
        analyses.append({'title': '绩效表现分析', 'type': 'info', 'content': perf_text})

    # 5. Age structure
    young = age_groups.get('25岁以下', 0) + age_groups.get('25-29岁', 0)
    middle = age_groups.get('30-34岁', 0) + age_groups.get('35-39岁', 0)
    senior = age_groups.get('40-44岁', 0) + age_groups.get('45-49岁', 0) + age_groups.get('50岁以上', 0)
    young_pct = round(young / total * 100, 1) if total else 0
    senior_pct = round(senior / total * 100, 1) if total else 0
    age_text = f'30岁以下 {young} 人（{young_pct}%），30-39岁 {middle} 人，40岁以上 {senior} 人（{senior_pct}%）。'
    if senior_pct > 50:
        age_text += f' 40岁以上员工占比 {senior_pct}%，年龄结构偏大，需加强年轻人才引进与梯队补充。'
        analyses.append({'title': '年龄结构分析', 'type': 'warning', 'content': age_text})
    elif young_pct > 35:
        age_text += f' 30岁以下员工占比 {young_pct}%，团队年轻化，富有活力，需注重经验传承与培养。'
        analyses.append({'title': '年龄结构分析', 'type': 'success', 'content': age_text})
    else:
        analyses.append({'title': '年龄结构分析', 'type': 'info', 'content': age_text})

    # 6. Education analysis
    bachelor_plus = edu_counts.get('本科', 0) + edu_counts.get('硕士研究生', 0) + edu_counts.get('硕士', 0) + edu_counts.get('博士', 0)
    high_school = edu_counts.get('高中/高职/中专', 0) + edu_counts.get('高中', 0)
    middle_school = edu_counts.get('初中', 0) + edu_counts.get('小学', 0)
    college = edu_counts.get('大专', 0)
    edu_pct = round(bachelor_plus / total * 100, 1) if total else 0
    low_edu_pct = round((high_school + middle_school) / total * 100, 1) if total else 0
    edu_text = f'本科及以上 {bachelor_plus} 人（{edu_pct}%），大专 {college} 人，高中及以下 {high_school + middle_school} 人（{low_edu_pct}%）。'
    if edu_pct < 20:
        edu_text += f' 本科及以上学历占比偏低（{edu_pct}%），建议针对核心岗位提升学历要求或加强专业培训。'
        analyses.append({'title': '学历结构分析', 'type': 'warning', 'content': edu_text})
    else:
        edu_text += f' 学历结构整体{("较好" if edu_pct > 40 else "适中")}。'
        analyses.append({'title': '学历结构分析', 'type': 'info', 'content': edu_text})

    # 7. Recommendations
    recommendations = []
    if star_count > 0:
        recommendations.append(f'- 超级明星员工 {star_count} 人是核心资产，建议制定个性化保留方案（晋升、轮岗、关键项目历练），防止人才流失。')
    if high_pot > 0:
        recommendations.append(f'- {high_pot} 名高潜人才需重点培养，建议安排导师辅导、跨部门轮岗、领导力培训等发展项目。')
    if problem_count > 0:
        recommendations.append(f'- {problem_count} 名问题/差距员工需启动绩效辅导，设定3-6个月改进期限，无明显改善则启动退出机制。')
    if c_count > 0:
        recommendations.append(f'- {c_count} 名C类绩效员工需一对一沟通，明确改进目标，必要时调整岗位或职责。')
    if senior_pct > 50:
        recommendations.append(f'- 团队年龄结构偏大，建议加大应届生/年轻人才招聘力度，做好知识沉淀与经验传承。')
    if t1 + t2 < total * 0.1:
        recommendations.append(f'- 关键梯队人才储备不足，建议从高潜员工中选拔建立后备人才库，制定加速发展计划。')
    if avg_match < 7:
        recommendations.append(f'- 平均人岗匹配评分 {avg_match} 分偏低，建议重新审视岗位配置与人岗匹配度，优化人员安排。')
    if not recommendations:
        recommendations.append('- 当前部门人才结构整体健康，建议持续关注人才发展与梯队建设，保持定期盘点与动态调整。')

    analyses.append({
        'title': '管理建议',
        'type': 'info',
        'content': '\n'.join(recommendations)
    })

    return analyses


@app.route('/api/export')
@login_required
def api_export():
    """Export filtered employee data and analysis as a beautifully formatted Excel file."""
    db = get_db()
    dept_ids = get_accessible_dept_ids()
    if not dept_ids:
        return jsonify({'error': '无可访问部门'}), 403

    category = request.args.get('category', '')
    sub_dept = request.args.get('sub_dept', '')
    grid_pos = request.args.get('grid', '')
    pipeline = request.args.get('pipeline', '')

    query = f"""
        SELECT e.*, d.category, d.sub_dept
        FROM employees e
        JOIN departments d ON e.dept_id = d.id
        WHERE e.dept_id IN ({','.join('?' * len(dept_ids))})
    """
    params = list(dept_ids)

    if category:
        query += ' AND d.category = ?'
        params.append(category)
    if sub_dept:
        query += ' AND d.sub_dept = ?'
        params.append(sub_dept)
    if grid_pos:
        query += ' AND e.grid_position = ?'
        params.append(int(grid_pos))
    if pipeline:
        query += ' AND e.talent_pipeline = ?'
        params.append(pipeline)

    query += ' ORDER BY d.category, d.sub_dept, e.id'
    rows = db.execute(query, params).fetchall()

    if not rows:
        return jsonify({'error': '无数据可导出'}), 404

    # ---- Compute statistics ----
    total = len(rows)
    grid_names = {k: v['name'] for k, v in GRID_INFO.items()}
    grid_colors = {k: v['color'] for k, v in GRID_INFO.items()}

    scope = ''
    if category:
        scope = category
    if sub_dept:
        scope += f' / {sub_dept}' if scope else sub_dept
    if not scope:
        scope = '全部可访问部门'
    if grid_pos:
        scope += f' (九宫格: {grid_names.get(int(grid_pos), grid_pos)})'
    if pipeline:
        scope += f' (梯队: {pipeline})'

    ages = [r['age'] for r in rows if r['age']]
    avg_age = round(sum(ages) / len(ages), 1) if ages else 0

    match_scores = [r['person_position_score'] for r in rows if r['person_position_score']]
    avg_match = round(sum(match_scores) / len(match_scores), 2) if match_scores else 0

    # Tenure calculation
    tenures = []
    for r in rows:
        t = r['company_tenure']
        if t and t != '':
            months = 0
            if '年' in t and '个月' in t:
                parts = t.split('年')
                months = int(parts[0]) * 12 if parts[0].strip() else 0
                rest = parts[1].replace('个月', '').strip()
                months += int(rest) if rest else 0
            elif '年' in t:
                months = int(t.replace('年', '').strip()) * 12
            elif '个月' in t:
                months = int(t.replace('个月', '').strip())
            if months > 0:
                tenures.append(months)
    avg_tenure = round(sum(tenures) / len(tenures), 1) if tenures else 0

    grid_counts = {}
    for r in rows:
        gp = r['grid_position']
        if gp:
            grid_counts[gp] = grid_counts.get(gp, 0) + 1

    pipeline_counts = {}
    for r in rows:
        p = r['talent_pipeline']
        if p:
            pipeline_counts[p] = pipeline_counts.get(p, 0) + 1

    perf_counts = {}
    for r in rows:
        p = r['annual_performance']
        if p:
            perf_counts[p] = perf_counts.get(p, 0) + 1

    edu_counts = {}
    for r in rows:
        e = r['education']
        if e:
            edu_counts[e] = edu_counts.get(e, 0) + 1

    age_groups = {'25岁以下': 0, '25-29岁': 0, '30-34岁': 0, '35-39岁': 0, '40-44岁': 0, '45-49岁': 0, '50岁以上': 0}
    for a in ages:
        if a < 25: age_groups['25岁以下'] += 1
        elif a < 30: age_groups['25-29岁'] += 1
        elif a < 35: age_groups['30-34岁'] += 1
        elif a < 40: age_groups['35-39岁'] += 1
        elif a < 45: age_groups['40-44岁'] += 1
        elif a < 50: age_groups['45-49岁'] += 1
        else: age_groups['50岁以上'] += 1

    star_count = grid_counts.get(9, 0)
    high_pot = star_count + grid_counts.get(7, 0) + grid_counts.get(4, 0)
    high_perf = star_count + grid_counts.get(8, 0) + grid_counts.get(6, 0)
    core_count = grid_counts.get(5, 0)
    problem_count = grid_counts.get(1, 0) + grid_counts.get(2, 0)

    # ---- Build Excel workbook ----
    wb = Workbook()

    # Style definitions
    header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Microsoft YaHei', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name='Microsoft YaHei', bold=True, size=16, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    subtitle_font = Font(name='Microsoft YaHei', size=10, color='808080')
    section_font = Font(name='Microsoft YaHei', bold=True, size=12, color='1F4E79')
    section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    alt_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
    metric_font = Font(name='Microsoft YaHei', bold=True, size=20, color='4472C4')
    metric_label_font = Font(name='Microsoft YaHei', size=9, color='808080')
    metric_fill = PatternFill(start_color='EAF2FA', end_color='EAF2FA', fill_type='solid')
    warning_font = Font(name='Microsoft YaHei', size=10, color='CC0000')
    success_font = Font(name='Microsoft YaHei', size=10, color='008000')
    info_font = Font(name='Microsoft YaHei', size=10, color='333333')

    # ================================================================
    # Sheet 1: Employee Data
    # ================================================================
    ws1 = wb.active
    ws1.title = '员工数据'

    headers = [
        '序号', '大类', '子部门', '中文姓名', '英文姓名', '岗位名称',
        '岗位职责', '职级', '年龄', '学历', '毕业学校', '毕业日期',
        '工作年限', '入职日期', '司龄',
        '知识技能匹配', '解决问题匹配', '职务责任匹配', '人岗匹配评分',
        '年度绩效', '学习能力', '思考能力', '理解他人', '情感成熟度',
        '潜力评分', '绩效等级', '潜力等级', '九宫格位置', '九宫格名称',
        '所在梯队', '结果运用', '发展计划', '管理策略'
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(rows, 2):
        emp = dict(r)
        values = [
            row_idx - 1,
            emp.get('category', ''),
            emp.get('sub_dept', ''),
            emp.get('chinese_name', ''),
            emp.get('english_name', ''),
            emp.get('position_title', ''),
            emp.get('job_responsibility', ''),
            emp.get('job_level', ''),
            emp.get('age', ''),
            emp.get('education', ''),
            emp.get('graduation_institution', ''),
            emp.get('graduation_date', ''),
            emp.get('work_experience', ''),
            emp.get('entry_date', ''),
            emp.get('company_tenure', ''),
            emp.get('knowledge_skill_match', ''),
            emp.get('problem_solving_match', ''),
            emp.get('responsibility_match', ''),
            emp.get('person_position_score', ''),
            emp.get('annual_performance', ''),
            emp.get('learning_ability', ''),
            emp.get('thinking_ability', ''),
            emp.get('understanding_others', ''),
            emp.get('emotional_maturity', ''),
            emp.get('potential_score', ''),
            emp.get('performance_level', ''),
            emp.get('potential_level', ''),
            emp.get('grid_position', ''),
            grid_names.get(emp.get('grid_position'), ''),
            emp.get('talent_pipeline', ''),
            emp.get('result_application', ''),
            emp.get('development_plan', ''),
            emp.get('management_strategy', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            if val is None:
                val = ''
            val = _sanitize_for_excel(val)
            ws1.cell(row=row_idx, column=col_idx, value=val)

    col_widths = [6, 8, 14, 10, 12, 16, 30, 8, 6, 10, 16, 12, 10, 12, 10,
                  10, 10, 10, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 10, 10, 20, 30, 20]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = 'A2'

    # ================================================================
    # Sheet 2: Analysis Report (with charts)
    # ================================================================
    ws2 = wb.create_sheet('分析报告')

    # Column widths
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12

    row_num = 1

    # --- Title ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='人才盘点分析报告')
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[row_num].height = 40
    row_num += 1

    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value=f'分析范围：{scope}    导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal='center')
    row_num += 2

    # --- Section 1: Overview ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='一、总体概况')
    cell.font = section_font
    cell.fill = section_fill
    row_num += 1

    # Metric cards (2 rows x 4 columns)
    metrics = [
        ('总人数', f'{total} 人'),
        ('平均年龄', f'{avg_age} 岁'),
        ('平均司龄', f'{avg_tenure} 月'),
        ('平均人岗匹配', f'{avg_match} 分'),
        ('高潜人才', f'{high_pot} 人'),
        ('高绩效人才', f'{high_perf} 人'),
        ('超级明星', f'{star_count} 人'),
        ('中坚力量', f'{core_count} 人'),
    ]
    for i in range(0, len(metrics), 3):
        batch = metrics[i:i+3]
        for j, (label, value) in enumerate(batch):
            col = j * 2 + 1
            ws2.merge_cells(start_row=row_num, start_column=col, end_row=row_num, end_column=col+1)
            c = ws2.cell(row=row_num, column=col, value=label)
            c.font = metric_label_font
            c.fill = metric_fill
            c.alignment = center_align
        row_num += 1
        for j, (label, value) in enumerate(batch):
            col = j * 2 + 1
            ws2.merge_cells(start_row=row_num, start_column=col, end_row=row_num, end_column=col+1)
            c = ws2.cell(row=row_num, column=col, value=value)
            c.font = metric_font
            c.fill = metric_fill
            c.alignment = center_align
        ws2.row_dimensions[row_num].height = 32
        row_num += 1
    row_num += 1

    # --- Section 2: Grid Distribution ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='二、九宫格分布')
    cell.font = section_font
    cell.fill = section_fill
    row_num += 1

    grid_header_row = row_num
    for c, h in enumerate(['位置', '名称', '人数', '占比'], 1):
        cell = ws2.cell(row=row_num, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    row_num += 1

    grid_data_start = row_num
    for gp in range(1, 10):
        info = GRID_INFO.get(gp, {})
        count = grid_counts.get(gp, 0)
        pct = f'{round(count / total * 100, 1)}%' if total else '0%'
        vals = [gp, info.get('name', ''), count, pct]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_num, column=c, value=v)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
            if count == 0:
                cell.font = Font(name='Microsoft YaHei', size=10, color='CCCCCC')
        row_num += 1
    grid_data_end = row_num - 1
    row_num += 1

    # Grid chart (BarChart) - placed below table
    chart_grid = BarChart()
    chart_grid.type = 'col'
    chart_grid.title = '九宫格分布'
    chart_grid.y_axis.title = '人数'
    chart_grid.x_axis.title = '位置'
    chart_grid.height = 7
    chart_grid.width = 16
    data_ref = Reference(ws2, min_col=3, min_row=grid_header_row, max_row=grid_data_end)
    cat_ref = Reference(ws2, min_col=2, min_row=grid_data_start, max_row=grid_data_end)
    chart_grid.add_data(data_ref, titles_from_data=True)
    chart_grid.set_categories(cat_ref)
    chart_grid.legend = None
    chart_grid.dataLabels = DataLabelList(showVal=True)
    ws2.add_chart(chart_grid, f'A{row_num}')
    row_num += 16

    # --- Section 3: Pipeline Distribution ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='三、人才梯队分布')
    cell.font = section_font
    cell.fill = section_fill
    row_num += 1

    pipe_header_row = row_num
    for c, h in enumerate(['梯队', '', '人数', '占比'], 1):
        cell = ws2.cell(row=row_num, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    row_num += 1

    pipe_data_start = row_num
    for p in PIPELINE_ORDER:
        count = pipeline_counts.get(p, 0)
        pct = f'{round(count / total * 100, 1)}%' if total else '0%'
        ws2.cell(row=row_num, column=1, value=p).font = cell_font
        ws2.cell(row=row_num, column=3, value=count).font = cell_font
        ws2.cell(row=row_num, column=4, value=pct).font = cell_font
        for c in range(1, 5):
            ws2.cell(row=row_num, column=c).alignment = center_align
            ws2.cell(row=row_num, column=c).border = thin_border
        row_num += 1
    pipe_data_end = row_num - 1
    row_num += 1

    chart_pipe = BarChart()
    chart_pipe.type = 'bar'
    chart_pipe.title = '人才梯队分布'
    chart_pipe.y_axis.title = '梯队'
    chart_pipe.x_axis.title = '人数'
    chart_pipe.height = 7
    chart_pipe.width = 16
    data_ref = Reference(ws2, min_col=3, min_row=pipe_header_row, max_row=pipe_data_end)
    cat_ref = Reference(ws2, min_col=1, min_row=pipe_data_start, max_row=pipe_data_end)
    chart_pipe.add_data(data_ref, titles_from_data=True)
    chart_pipe.set_categories(cat_ref)
    chart_pipe.legend = None
    chart_pipe.dataLabels = DataLabelList(showVal=True)
    ws2.add_chart(chart_pipe, f'A{row_num}')
    row_num += 16

    # --- Section 4: Performance Distribution ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='四、绩效表现分布')
    cell.font = section_font
    cell.fill = section_fill
    row_num += 1

    perf_header_row = row_num
    for c, h in enumerate(['等级', '说明', '人数', '占比'], 1):
        cell = ws2.cell(row=row_num, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    row_num += 1

    perf_data_start = row_num
    perf_labels = {'A': '优秀', 'B': '达标', 'C': '待改进'}
    for p in ['A', 'B', 'C']:
        count = perf_counts.get(p, 0)
        pct = f'{round(count / total * 100, 1)}%' if total else '0%'
        ws2.cell(row=row_num, column=1, value=p).font = cell_font
        ws2.cell(row=row_num, column=2, value=perf_labels.get(p, '')).font = cell_font
        ws2.cell(row=row_num, column=3, value=count).font = cell_font
        ws2.cell(row=row_num, column=4, value=pct).font = cell_font
        for c in range(1, 5):
            ws2.cell(row=row_num, column=c).alignment = center_align
            ws2.cell(row=row_num, column=c).border = thin_border
        row_num += 1
    perf_data_end = row_num - 1
    row_num += 1

    chart_perf = PieChart()
    chart_perf.title = '绩效表现分布'
    chart_perf.height = 7
    chart_perf.width = 12
    data_ref = Reference(ws2, min_col=3, min_row=perf_header_row, max_row=perf_data_end)
    cat_ref = Reference(ws2, min_col=1, min_row=perf_data_start, max_row=perf_data_end)
    chart_perf.add_data(data_ref, titles_from_data=True)
    chart_perf.set_categories(cat_ref)
    chart_perf.legend = None
    chart_perf.dataLabels = DataLabelList(showVal=True, showPercent=True)
    ws2.add_chart(chart_perf, f'A{row_num}')
    row_num += 16

    # --- Section 5: Education Distribution ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='五、学历分布')
    cell.font = section_font
    cell.fill = section_fill
    row_num += 1

    edu_header_row = row_num
    for c, h in enumerate(['学历', '', '人数', '占比'], 1):
        cell = ws2.cell(row=row_num, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    row_num += 1

    edu_data_start = row_num
    edu_order = ['硕士研究生', '硕士', '博士', '本科', '大专', '高中/高职/中专', '高中', '初中', '小学']
    sorted_edu = sorted(edu_counts.items(), key=lambda x: (edu_order.index(x[0]) if x[0] in edu_order else 99, -x[1]))
    for edu, count in sorted_edu:
        pct = f'{round(count / total * 100, 1)}%' if total else '0%'
        ws2.cell(row=row_num, column=1, value=edu).font = cell_font
        ws2.cell(row=row_num, column=3, value=count).font = cell_font
        ws2.cell(row=row_num, column=4, value=pct).font = cell_font
        for c in range(1, 5):
            ws2.cell(row=row_num, column=c).alignment = center_align
            ws2.cell(row=row_num, column=c).border = thin_border
        row_num += 1
    edu_data_end = row_num - 1
    row_num += 1

    if sorted_edu:
        chart_edu = PieChart()
        chart_edu.title = '学历分布'
        chart_edu.height = 7
        chart_edu.width = 12
        data_ref = Reference(ws2, min_col=3, min_row=edu_header_row, max_row=edu_data_end)
        cat_ref = Reference(ws2, min_col=1, min_row=edu_data_start, max_row=edu_data_end)
        chart_edu.add_data(data_ref, titles_from_data=True)
        chart_edu.set_categories(cat_ref)
        chart_edu.legend = None
        chart_edu.dataLabels = DataLabelList(showVal=True, showPercent=True)
        ws2.add_chart(chart_edu, f'A{row_num}')
        row_num += 16

    # --- Section 6: Age Distribution ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='六、年龄分布')
    cell.font = section_font
    cell.fill = section_fill
    row_num += 1

    age_header_row = row_num
    for c, h in enumerate(['年龄段', '', '人数', '占比'], 1):
        cell = ws2.cell(row=row_num, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    row_num += 1

    age_data_start = row_num
    for age_label, count in age_groups.items():
        pct = f'{round(count / total * 100, 1)}%' if total else '0%'
        ws2.cell(row=row_num, column=1, value=age_label).font = cell_font
        ws2.cell(row=row_num, column=3, value=count).font = cell_font
        ws2.cell(row=row_num, column=4, value=pct).font = cell_font
        for c in range(1, 5):
            ws2.cell(row=row_num, column=c).alignment = center_align
            ws2.cell(row=row_num, column=c).border = thin_border
        row_num += 1
    age_data_end = row_num - 1
    row_num += 1

    chart_age = BarChart()
    chart_age.type = 'col'
    chart_age.title = '年龄分布'
    chart_age.y_axis.title = '人数'
    chart_age.x_axis.title = '年龄段'
    chart_age.height = 7
    chart_age.width = 16
    data_ref = Reference(ws2, min_col=3, min_row=age_header_row, max_row=age_data_end)
    cat_ref = Reference(ws2, min_col=1, min_row=age_data_start, max_row=age_data_end)
    chart_age.add_data(data_ref, titles_from_data=True)
    chart_age.set_categories(cat_ref)
    chart_age.legend = None
    chart_age.dataLabels = DataLabelList(showVal=True)
    ws2.add_chart(chart_age, f'A{row_num}')
    row_num += 16

    # --- Section 7: Key Talent ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='七、关键人才统计')
    cell.font = section_font
    cell.fill = section_fill
    row_num += 1

    key_items = [
        ('超级明星 (九宫格9)', star_count, '高绩效+高潜力，核心资产'),
        ('高潜人才 (4+7+9)', high_pot, '需重点培养与发展'),
        ('高绩效人才 (6+8+9)', high_perf, '可依靠的稳定贡献者'),
        ('中坚力量 (5)', core_count, '稳定中坚，开发培训提升'),
        ('问题/差距员工 (1+2)', problem_count, '需绩效辅导或退出'),
    ]
    for c, h in enumerate(['类别', '', '人数', '占比', '说明'], 1):
        cell = ws2.cell(row=row_num, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    row_num += 1
    for label, count, desc in key_items:
        pct = f'{round(count / total * 100, 1)}%' if total else '0%'
        ws2.cell(row=row_num, column=1, value=label).font = cell_font
        ws2.cell(row=row_num, column=3, value=count).font = Font(name='Microsoft YaHei', size=10, bold=True)
        ws2.cell(row=row_num, column=4, value=pct).font = cell_font
        ws2.cell(row=row_num, column=5, value=desc).font = Font(name='Microsoft YaHei', size=9, color='666666')
        for c in range(1, 6):
            ws2.cell(row=row_num, column=c).alignment = center_align
            ws2.cell(row=row_num, column=c).border = thin_border
        row_num += 1
    row_num += 1

    # --- Section 8: Intelligent Analysis ---
    ws2.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws2.cell(row=row_num, column=1, value='八、智能分析与建议')
    cell.font = section_font
    cell.fill = section_fill
    row_num += 1

    # Generate analysis text
    analyses = _generate_analysis_text(total, scope, avg_age, avg_tenure, avg_match,
                                       grid_counts, pipeline_counts, perf_counts,
                                       edu_counts, age_groups, star_count, high_pot,
                                       high_perf, core_count, problem_count)
    for a in analyses:
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        cell = ws2.cell(row=row_num, column=1, value=f'【{a["title"]}】')
        cell.font = Font(name='Microsoft YaHei', bold=True, size=10, color='2E75B6')
        row_num += 1

        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        cell = ws2.cell(row=row_num, column=1, value=a['content'])
        if a['type'] == 'warning':
            cell.font = warning_font
        elif a['type'] == 'success':
            cell.font = success_font
        else:
            cell.font = info_font
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws2.row_dimensions[row_num].height = 45
        row_num += 1

    # ================================================================
    # Generate filename & return
    # ================================================================
    filename_parts = ['人才盘点']
    if category:
        filename_parts.append(category)
    if sub_dept:
        filename_parts.append(sub_dept)
    if grid_pos:
        filename_parts.append(f'九宫格{grid_pos}')
    if pipeline:
        filename_parts.append(pipeline)
    filename_parts.append(datetime.now().strftime('%Y%m%d'))
    filename = '_'.join(filename_parts) + '.xlsx'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    encoded_filename = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ============================================================
# Main
# ============================================================

# Initialize database on import (required for gunicorn/production)
init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    print(f"Database: {DB_PATH}")
    print(f"Starting server at http://0.0.0.0:{port}")
    print("\nDefault accounts:")
    print("  HR:      username=hr_admin  password=hr123456")
    print("  部门管理员: admin_AT(印染) / admin_GB(服装) / admin_KN(针织) / admin_FL(辅料) / admin_ZB(总部)")
    app.run(host='0.0.0.0', port=port, debug=False)
